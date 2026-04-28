"""
MDB Employee Exporter
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Reads your ZKTeco .mdb database and exports a clean
employee list with:
  Badgenumber | Name | DEPTNAME

Excel structure (two-level grouping):
  ┌─ ADMIN
  │     employee rows (DEPTNAME = ADMIN)
  ├─ SUPPORT
  │     employee rows (DEPTNAME = SUPPORT)
  ├─ TEACHING
  │     employee rows (DEPTNAME = TEACHING)
  ├─ TRANSPORT DEPT
  │   ├─ CONDUCTOR  →  employee rows
  │   └─ DRIVER     →  employee rows
  ├─ CLEANING STAFF
  │     employee rows (DEPTNAME = CLEANING STAFF)
  ├─ DELETED
  │   ├─ DELETED EMPLOYEES  →  employee rows
  │   └─ TRANSPORT          →  employee rows
  └─ GAES
      ├─ GAES                      →  employee rows
      └─ GULF ASIAN ENGLISH SCHOOL →  employee rows

Usage:
  python EXPORT_EMPLOYEES.py                      # auto-finds .mdb in same folder
  python EXPORT_EMPLOYEES.py path\\to\\backup.mdb # specify path

Output:
  employees_export.xlsx  +  employees_export.csv
  Saved in the same folder as this script.

Requirements:
  pip install pyodbc openpyxl pandas
  Install Access driver if needed:
  https://www.microsoft.com/en-us/download/details.aspx?id=54920
"""

import sys
import os
import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════════════
#  GROUPING CONFIG
#  Each entry: (PARENT_SECTION, [list of raw dept names that belong to it])
#  Depts with no parent (standalone) have their own name as parent.
#  Order here = order in the output file.
# ══════════════════════════════════════════════════════════════════════════════
SECTION_CONFIG = [
    ("ADMIN",          ["ADMIN"]),
    ("SUPPORT",        ["SUPPORT"]),
    ("TEACHING",       ["TEACHING"]),
    ("TRANSPORT DEPT", ["CONDUCTOR", "DRIVER"]),
    ("CLEANING STAFF", ["CLEANING STAFF"]),
    ("DELETED",        ["DELETED EMPLOYEES", "TRANSPORT"]),
    ("GAES",           ["GAES", "GULF ASIAN ENGLISH SCHOOL"]),
]

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def find_mdb():
    if len(sys.argv) > 1:
        path = sys.argv[1]
        if not os.path.exists(path):
            print(f"[ERROR] File not found: {path}")
            sys.exit(1)
        return path
    folder = os.path.dirname(os.path.abspath(__file__))
    for f in os.listdir(folder):
        if f.lower().endswith(".mdb") or f.lower().endswith(".accdb"):
            return os.path.join(folder, f)
    print("[ERROR] No .mdb / .accdb file found.")
    print("  Place the database in the same folder as this script, or pass the path:")
    print("  python EXPORT_EMPLOYEES.py path\\to\\backup.mdb")
    input("\nPress Enter to exit...")
    sys.exit(1)


def connect_db(mdb_path):
    try:
        return pyodbc.connect(
            r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
            f"Dbq={mdb_path};"
        )
    except Exception as e:
        print(f"[ERROR] Cannot open database: {e}")
        print("Install the Access driver:")
        print("  https://www.microsoft.com/en-us/download/details.aspx?id=54920")
        input("\nPress Enter to exit...")
        sys.exit(1)


# ── Excel styling ─────────────────────────────────────────────────────────────
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_col(ws, row, col, val):
    """Blue column header (Badgenumber / Name / Department)."""
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _thin()

def _section_hdr(ws, row, label):
    """Dark section header row — parent group (e.g. TRANSPORT DEPT)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = _thin()
    ws.row_dimensions[row].height = 20

def _sub_hdr(ws, row, label):
    """Light sub-department header row (e.g. CONDUCTOR, DRIVER)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name="Arial", bold=True, color="1F4E79", size=10)
    c.fill      = PatternFill("solid", start_color="DCE6F1")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border    = _thin()
    ws.row_dimensions[row].height = 17

def _data_cell(ws, row, col, val, bg, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _thin()
    ws.row_dimensions[row].height = 16


# ══════════════════════════════════════════════════════════════════════════════
#  BUILD SORT KEYS
# ══════════════════════════════════════════════════════════════════════════════
def build_sort_keys():
    """
    Returns two dicts:
      dept_to_section  : raw dept name -> parent section name
      dept_sort_order  : raw dept name -> (section_idx, dept_idx)
    """
    dept_to_section = {}
    dept_sort_order = {}
    for sec_idx, (section, depts) in enumerate(SECTION_CONFIG):
        for dept_idx, dept in enumerate(depts):
            dept_upper = dept.upper()
            dept_to_section[dept_upper] = section
            dept_sort_order[dept_upper] = (sec_idx, dept_idx)
    return dept_to_section, dept_sort_order


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    mdb_path = find_mdb()
    print(f"\n{'='*55}")
    print(f"  MDB Employee Exporter")
    print(f"{'='*55}")
    print(f"  Source : {mdb_path}")

    conn = connect_db(mdb_path)

    # ── Step 1: Load DEPARTMENTS ───────────────────────────────────────────────
    print("\n[1/4] Loading DEPARTMENTS table...")
    try:
        dept_df = pd.read_sql("SELECT [DEPTID], [DEPTNAME] FROM [DEPARTMENTS]", conn)
    except Exception as e:
        print(f"[ERROR] Could not read DEPARTMENTS: {e}")
        input("\nPress Enter to exit..."); sys.exit(1)

    # DEPTID (str) -> DEPTNAME (uppercase)
    id_to_name = dict(zip(
        dept_df["DEPTID"].astype(str).str.strip(),
        dept_df["DEPTNAME"].astype(str).str.strip().str.upper()
    ))
    print(f"  Found {len(id_to_name)} departments")

    # ── Step 2: Load USERINFO ──────────────────────────────────────────────────
    print("\n[2/4] Loading USERINFO table...")
    try:
        user_df = pd.read_sql(
            "SELECT [Badgenumber], [Name], [DEFAULTDEPTID] FROM [USERINFO]", conn
        )
    except Exception as e:
        print(f"[ERROR] Could not read USERINFO: {e}")
        input("\nPress Enter to exit..."); sys.exit(1)
    conn.close()
    print(f"  Found {len(user_df)} employees")

    # ── Step 3: Map, sort, group ───────────────────────────────────────────────
    print("\n[3/4] Mapping and sorting...")
    user_df["Badgenumber"]   = user_df["Badgenumber"].astype(str).str.strip()
    user_df["Name"]          = user_df["Name"].astype(str).str.strip()
    user_df["DEFAULTDEPTID"] = user_df["DEFAULTDEPTID"].astype(str).str.strip()

    # Resolve dept name from ID
    user_df["DEPTNAME"] = user_df["DEFAULTDEPTID"].map(id_to_name).fillna("UNKNOWN")

    dept_to_section, dept_sort_order = build_sort_keys()

    # Section name for grouping header
    user_df["_SECTION"] = user_df["DEPTNAME"].map(dept_to_section).fillna("UNKNOWN")

    # Sort key: (section_idx, dept_idx, name)
    user_df["_SORT"] = user_df["DEPTNAME"].map(
        lambda d: dept_sort_order.get(d.upper(), (999, 999))
    )
    user_df["_SEC_IDX"]  = user_df["_SORT"].map(lambda x: x[0])
    user_df["_DEPT_IDX"] = user_df["_SORT"].map(lambda x: x[1])
    user_df = user_df.sort_values(["_SEC_IDX", "_DEPT_IDX", "Name"]).reset_index(drop=True)

    out_df = user_df[["Badgenumber", "Name", "DEPTNAME", "_SECTION"]].copy()

    # Console summary
    print("\n  Breakdown:")
    for section, depts in SECTION_CONFIG:
        sec_total = 0
        for dept in depts:
            count = len(out_df[out_df["DEPTNAME"] == dept.upper()])
            sec_total += count
            if len(depts) > 1:
                print(f"    {section:20s}  ↳ {dept:30s} {count:4d}")
            else:
                print(f"    {section:20s}                               {count:4d}")
        if len(depts) > 1:
            print(f"    {'':20s}    {'SUBTOTAL':30s} {sec_total:4d}")
    print(f"\n  Total: {len(out_df)} employees")

    # ── Step 4: Save ───────────────────────────────────────────────────────────
    print("\n[4/4] Saving files...")
    out_dir   = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(out_dir, "employees_export.xlsx")
    csv_path  = os.path.join(out_dir, "employees_export.csv")

    # CSV — flat, keep original dept name
    out_df[["Badgenumber", "Name", "DEPTNAME"]].to_csv(
        csv_path, index=False, encoding="utf-8-sig"
    )
    print(f"  OK CSV  -> {csv_path}")

    # ── Excel ──────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "EMPLOYEES"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1,
                value=f"Employee Export  —  {datetime.now().strftime('%d %b %Y %H:%M')}")
    c.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    for col, hdr in enumerate(["Badgenumber", "Name", "Department"], 1):
        _hdr_col(ws, 2, col, hdr)
    ws.row_dimensions[2].height = 20

    # Write rows section by section
    xl_row      = 3
    prev_section = None
    prev_dept    = None
    alt          = 0

    for _, rec in out_df.iterrows():
        section = rec["_SECTION"]
        dept    = rec["DEPTNAME"]

        # New section header
        if section != prev_section:
            _section_hdr(ws, xl_row, section)
            xl_row      += 1
            prev_section = section
            prev_dept    = None
            alt          = 0

        # New sub-dept header (only when section has multiple depts)
        section_depts = next(d for s, d in SECTION_CONFIG if s == section)
        if len(section_depts) > 1 and dept != prev_dept:
            _sub_hdr(ws, xl_row, dept)
            xl_row   += 1
            prev_dept = dept
            alt       = 0

        # Data row
        bg = "EBF3FB" if alt % 2 == 0 else "FFFFFF"
        _data_cell(ws, xl_row, 1, rec["Badgenumber"], bg, align="center")
        _data_cell(ws, xl_row, 2, rec["Name"],        bg)
        _data_cell(ws, xl_row, 3, rec["DEPTNAME"],    bg)
        xl_row += 1
        alt    += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 28

    wb.save(xlsx_path)
    print(f"  OK XLSX -> {xlsx_path}")

    print(f"\n{'='*55}")
    print(f"  Done! {len(out_df)} employees exported.")
    print(f"{'='*55}")
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
